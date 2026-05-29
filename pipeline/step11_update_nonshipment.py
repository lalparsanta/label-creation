"""
Step 11 — Update Non-Shipment Tracker
  - After DHL labels are created, record every labeled order in the
    non-shipment Excel file (OIDS sheet) so the next day's pipeline
    (step2_filter_existing) skips them automatically.
  - Splits merged order IDs (e.g. "ID1, ID2" from same-buyer merge in step 8)
    into individual rows.
  - Skips any order ID already present in the tracker — no duplicates.
"""

import os
from datetime import date

import pandas as pd
import openpyxl

from pipeline.pipeline_state import PipelineState

OIDS_SHEET      = "OIDS"
SHIPPED_STATUS  = "SHIPPED - DHL"
AMAZON_URL      = "https://sellercentral.amazon.co.uk/orders-v3/order/{oid}"


# ── Helpers ───────────────────────────────────────────────────────────────────

def _load_existing_oids(filepath: str) -> set[str]:
    """Return the set of OIDs already tracked (normalised upper-case)."""
    df = pd.read_excel(filepath, sheet_name=OIDS_SHEET, dtype=str)
    if "OID" not in df.columns:
        return set()
    return set(df["OID"].dropna().str.strip().str.upper())


def _append_rows(filepath: str, new_rows: list[dict]) -> None:
    """Append rows to the OIDS sheet, preserving existing data and formatting."""
    wb = openpyxl.load_workbook(filepath)
    ws = wb[OIDS_SHEET]
    next_row = ws.max_row + 1
    for row in new_rows:
        ws.cell(row=next_row, column=1, value=row["DATE"])
        ws.cell(row=next_row, column=2, value=row["URL"])
        ws.cell(row=next_row, column=3, value=row["OID"])
        ws.cell(row=next_row, column=4, value=row["STATUS"])
        ws.cell(row=next_row, column=5, value=row.get("AMZ STATUS", ""))
        next_row += 1
    wb.save(filepath)
    wb.close()


# ── Step runner ───────────────────────────────────────────────────────────────

def run(state: PipelineState, non_shipped_filepath: str | None = None) -> PipelineState:
    """
    Appends labeled orders to the non-shipment tracker Excel file.

    Args:
        state:                  current PipelineState (post step 10)
        non_shipped_filepath:   path to non-shipment.xlsx — same file used in
                                step 2 to filter existing orders.
    """
    print("\n[Step 11] Updating non-shipment tracker")

    if not non_shipped_filepath or not os.path.exists(non_shipped_filepath):
        print("  No non-shipment file provided or file not found — skipping.")
        state.log("step11_update_nonshipment", len(state.df), "Skipped — no file provided")
        return state

    # ── Extract all individual order IDs from the pipeline state ─────────────
    # Step 8 may have merged same-buyer orders into comma-separated IDs in one cell.
    raw_ids: pd.Series = state.df["order-id"].dropna().astype(str)
    order_ids: list[str] = []
    for raw in raw_ids:
        for oid in raw.split(","):
            oid = oid.strip()
            if oid:
                order_ids.append(oid)

    print(f"  Orders to record: {len(order_ids)}")

    # ── Load existing tracker to avoid duplicates ─────────────────────────────
    try:
        existing_oids = _load_existing_oids(non_shipped_filepath)
    except Exception as e:
        print(f"  WARNING: Could not read non-shipment file ({e}) — skipping.")
        state.log("step11_update_nonshipment", len(state.df), f"Skipped — read error: {e}")
        return state

    print(f"  Already in tracker: {len(existing_oids)}")

    today_date = date.today().strftime("%d-%b-%Y").upper()   # e.g. 25-MAY-2026

    new_rows: list[dict] = []
    skipped: list[str] = []

    for oid in order_ids:
        if oid.upper() in existing_oids:
            skipped.append(oid)
        else:
            new_rows.append({
                "DATE":       today_date,
                "URL":        AMAZON_URL.format(oid=oid),
                "OID":        oid,
                "STATUS":     SHIPPED_STATUS,
                "AMZ STATUS": "",
            })
            existing_oids.add(oid.upper())   # prevent duplicates within this batch

    if skipped:
        print(f"  Already tracked — skipping {len(skipped)} order(s)")

    if not new_rows:
        print("  Nothing new to add — tracker already up to date.")
        state.log("step11_update_nonshipment", len(state.df), "No new orders to add")
        return state

    # ── Write to Excel ────────────────────────────────────────────────────────
    try:
        _append_rows(non_shipped_filepath, new_rows)
    except Exception as e:
        print(f"  WARNING: Could not write to non-shipment file ({e})")
        print("  Make sure the file is not open in Excel.")
        state.log("step11_update_nonshipment", len(state.df), f"Write error: {e}")
        return state

    print(f"  Added {len(new_rows)} order(s) with status '{SHIPPED_STATUS}'")

    state.log(
        "step11_update_nonshipment",
        len(state.df),
        f"Added {len(new_rows)} orders to non-shipment tracker",
    )
    return state
