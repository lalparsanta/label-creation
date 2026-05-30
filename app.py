"""
DHL Label Creator — Streamlit Web App
Upload Amazon unshipped .txt → get DHL import CSV in your browser.

Persistence:
  - Weight files and D-zone are bundled in the repo (reference data/).
  - Non-shipment tracker is read from and written back to the private GitHub
    repo automatically via the GitHub API (requires GITHUB_TOKEN + GITHUB_REPO
    Streamlit secrets). Falls back to the bundled file if secrets are absent.
"""

import streamlit as st
import pandas as pd
import io
import os
import sys
import base64
import json
import urllib.request
import urllib.parse
import tempfile
import shutil
import threading
import contextlib
from datetime import date

# ─── Path setup (must run before pipeline imports) ────────────────────────────
ROOT = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, ROOT)

# ─── Pipeline module imports (for patching module-level constants) ────────────
import pipeline.step4_sku_gate as _step4_mod
import pipeline.step7_dzone as _step7_mod
import pipeline.step10_dhl_export as _step10_mod

from pipeline.step1_ingest import run as _step1
from pipeline.step2_filter_existing import run as _step2
from pipeline.step3_parse_sku import run as _step3
from pipeline.step5_weight import run as _step5
from pipeline.step6_validate import run as _step6
from pipeline.step7_dzone import run as _step7
from pipeline.step8_address import run as _step8
from pipeline.step9_instructions import run as _step9
from pipeline.step10_dhl_export import run as _step10
from pipeline.step11_update_nonshipment import run as _step11

# ─── Bundled default reference file paths ────────────────────────────────────
_REF          = os.path.join(ROOT, "reference data")
_DEFAULT_NS   = os.path.join(_REF, "non-shipment.xlsx")
_DEFAULT_PNC  = os.path.join(_REF, "PNC & BTC WEIGHT FILE.xlsx")
_DEFAULT_DFF  = os.path.join(_REF, "DFF WEIGHT FILE.xlsx")
_DEFAULT_DZONE = os.path.join(_REF, "D ZONE AREAS 1.xlsx")

# GitHub paths within the repo (must match where the files actually live)
_GH_NS_PATH   = "reference data/non-shipment.xlsx"

_pipeline_lock = threading.Lock()


# ─── Page config ──────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="DHL Label Creator",
    page_icon="📦",
    layout="wide",
)


# ─── GitHub API helpers ───────────────────────────────────────────────────────
def _gh_get(token: str, repo: str, path: str) -> tuple[bytes | None, str | None]:
    """Download a file from a private GitHub repo. Returns (content_bytes, sha)."""
    url = f"https://api.github.com/repos/{repo}/contents/{urllib.parse.quote(path)}"
    req = urllib.request.Request(url, headers={
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "User-Agent": "DHL-Label-Creator",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = json.loads(resp.read())
        return base64.b64decode(data["content"]), data["sha"]
    except Exception:
        return None, None


def _gh_put(token: str, repo: str, path: str,
            content: bytes, sha: str, message: str) -> str | None:
    """Commit a file update to GitHub. Returns the new SHA, or None on failure."""
    url = f"https://api.github.com/repos/{repo}/contents/{urllib.parse.quote(path)}"
    payload = json.dumps({
        "message": message,
        "content": base64.b64encode(content).decode("utf-8"),
        "sha": sha,
    }).encode("utf-8")
    req = urllib.request.Request(url, data=payload, method="PUT", headers={
        "Authorization": f"token {token}",
        "Accept": "application/vnd.github.v3+json",
        "Content-Type": "application/json",
        "User-Agent": "DHL-Label-Creator",
    })
    try:
        with urllib.request.urlopen(req, timeout=15) as resp:
            result = json.loads(resp.read())
        return result["content"]["sha"]
    except Exception:
        return None


def _gh_secrets() -> tuple[str, str]:
    """Return (token, repo) from Streamlit secrets, or ('', '') if not set."""
    try:
        return st.secrets.get("GITHUB_TOKEN", ""), st.secrets.get("GITHUB_REPO", "")
    except Exception:
        return "", ""


# ─── Session state init ───────────────────────────────────────────────────────
_DEFAULTS = {
    "stage":              "upload",
    "pipeline_state":     None,
    "logs_phase1":        "",
    "logs_phase2":        "",
    # Reference file bytes held in session — loaded once at startup
    "non_shipment_bytes": None,
    "ns_sha":             None,      # GitHub SHA needed to commit updates
    "ns_source":          None,      # "github" | "local" | None
    "pnc_btc_bytes":      None,
    "dff_bytes":          None,
    "dzone_bytes":        None,
    # Outputs
    "out_dhl":            None,
    "out_audit":          None,
    "out_nonship":        None,
    "out_dzone":          None,
    "gh_update_ok":       None,      # True/False/None — result of last GitHub commit
    # Settings
    "openai_key":         "",
    "use_ai":             False,
    "error":              None,
    # Tracking tab
    "tracking_out":       None,
    "tracking_log":       "",
    # Internal flags
    "_refs_loaded":       False,
}
for _k, _v in _DEFAULTS.items():
    if _k not in st.session_state:
        st.session_state[_k] = _v


# ─── Load reference files once per session ────────────────────────────────────
if not st.session_state._refs_loaded:
    token, repo = _gh_secrets()

    # Non-shipment: try GitHub first, then bundled local copy
    if token and repo:
        _ns_bytes, _ns_sha = _gh_get(token, repo, _GH_NS_PATH)
        if _ns_bytes:
            st.session_state.non_shipment_bytes = _ns_bytes
            st.session_state.ns_sha             = _ns_sha
            st.session_state.ns_source          = "github"
        else:
            st.session_state.ns_source = "github_error"
    elif os.path.exists(_DEFAULT_NS):
        with open(_DEFAULT_NS, "rb") as _f:
            st.session_state.non_shipment_bytes = _f.read()
        st.session_state.ns_source = "local"
    else:
        st.session_state.ns_source = None

    # Other reference files — load from disk once (they're bundled in the repo)
    for _attr, _path in [
        ("pnc_btc_bytes", _DEFAULT_PNC),
        ("dff_bytes",     _DEFAULT_DFF),
        ("dzone_bytes",   _DEFAULT_DZONE),
    ]:
        if os.path.exists(_path):
            with open(_path, "rb") as _f:
                st.session_state[_attr] = _f.read()

    st.session_state._refs_loaded = True


# ─── Pipeline helpers ─────────────────────────────────────────────────────────
def _write(data: bytes | None, directory: str, filename: str) -> str | None:
    if not data:
        return None
    path = os.path.join(directory, filename)
    with open(path, "wb") as f:
        f.write(data)
    return path


def _read_bytes(path: str | None) -> bytes | None:
    if path and os.path.exists(path):
        with open(path, "rb") as f:
            return f.read()
    return None


def _log_skipped_to_nonshipment(ns_path: str, skipped_oids: list[str]) -> None:
    """Append manually-removed order IDs to the non-shipment tracker as SKIPPED - MANUAL."""
    if not skipped_oids or not ns_path or not os.path.exists(ns_path):
        return
    import openpyxl
    OIDS_SHEET  = "OIDS"
    AMAZON_URL  = "https://sellercentral.amazon.co.uk/orders-v3/order/{oid}"
    today_str   = date.today().strftime("%d-%b-%Y").upper()
    wb = openpyxl.load_workbook(ns_path)
    if OIDS_SHEET not in wb.sheetnames:
        wb.close()
        return
    ws = wb[OIDS_SHEET]
    existing = {str(row[2]).strip().upper() for row in ws.iter_rows(min_row=2, values_only=True) if row[2]}
    next_row = ws.max_row + 1
    for oid in skipped_oids:
        if oid.strip().upper() not in existing:
            ws.cell(row=next_row, column=1, value=today_str)
            ws.cell(row=next_row, column=2, value=AMAZON_URL.format(oid=oid))
            ws.cell(row=next_row, column=3, value=oid)
            ws.cell(row=next_row, column=4, value="SKIPPED - MANUAL")
            ws.cell(row=next_row, column=5, value="")
            existing.add(oid.strip().upper())
            next_row += 1
    wb.save(ns_path)
    wb.close()


def _reset():
    for k in ("stage", "pipeline_state", "logs_phase1", "logs_phase2",
               "out_dhl", "out_audit", "out_nonship", "out_dzone",
               "gh_update_ok", "error"):
        st.session_state[k] = _DEFAULTS[k]


# ─── Phase 1: steps 1–3 ───────────────────────────────────────────────────────
def run_phase1(input_bytes: bytes) -> tuple:
    log_buf = io.StringIO()
    with _pipeline_lock:
        tmpdir = tempfile.mkdtemp()
        try:
            out_dir = os.path.join(tmpdir, "output")
            os.makedirs(out_dir)
            _step4_mod.PENDING_FILE = os.path.join(out_dir, "pending_sku_review.csv")

            input_path    = _write(input_bytes, tmpdir, "amazon_input.txt")
            non_ship_path = _write(st.session_state.non_shipment_bytes, tmpdir, "non-shipment.xlsx")

            with contextlib.redirect_stdout(log_buf):
                state = _step1(input_path)
                state = _step2(state, non_shipped_filepath=non_ship_path)
                state = _step3(state)

            return state, log_buf.getvalue()
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)


# ─── Phase 2: steps 5–11 ──────────────────────────────────────────────────────
def run_phase2(pipeline_state, skipped_oids: list[str] | None = None) -> tuple:
    log_buf = io.StringIO()
    outputs: dict[str, bytes] = {}

    with _pipeline_lock:
        tmpdir = tempfile.mkdtemp()
        try:
            out_dir = os.path.join(tmpdir, "output")
            os.makedirs(out_dir)

            _step7_mod.DZONE_DROPPED_FILE = os.path.join(out_dir, "dropped_dzone_orders.csv")
            _step10_mod.OUTPUT_DIR        = out_dir

            pnc_path  = _write(st.session_state.pnc_btc_bytes, tmpdir, "pnc_btc.xlsx")
            dff_path  = _write(st.session_state.dff_bytes,     tmpdir, "dff.xlsx")
            dzone_path = _write(st.session_state.dzone_bytes,  tmpdir, "dzone.xlsx")
            ns_path   = _write(st.session_state.non_shipment_bytes, tmpdir, "non-shipment.xlsx")

            openai_key = st.session_state.openai_key or None
            use_ai     = st.session_state.use_ai and bool(openai_key)

            with contextlib.redirect_stdout(log_buf):
                state = _step5(pipeline_state,
                               pnc_btc_weight_filepath=pnc_path,
                               dff_weight_filepath=dff_path)
                state = _step6(state)
                state = _step7(state, dzone_reference_filepath=dzone_path)
                state = _step8(state)
                state = _step9(state, openai_api_key=openai_key, use_ai=use_ai)
                _step10(state)
                _step11(state, non_shipped_filepath=ns_path)

            # Log manually-skipped orders so they're filtered next run
            if skipped_oids:
                _log_skipped_to_nonshipment(ns_path, skipped_oids)

            today = date.today().strftime("%Y-%m-%d")
            for fname, key in [
                (f"dhl_import_{today}.csv",  "dhl"),
                (f"audit_log_{today}.csv",   "audit"),
                ("dropped_dzone_orders.csv", "dzone"),
            ]:
                data = _read_bytes(os.path.join(out_dir, fname))
                if data:
                    outputs[key] = data

            if ns_path:
                data = _read_bytes(ns_path)
                if data:
                    outputs["non_shipment"] = data

            return outputs, log_buf.getvalue()
        finally:
            shutil.rmtree(tmpdir, ignore_errors=True)


# ─── Tracking tab helper ─────────────────────────────────────────────────────
def _run_tracking(csv_bytes: bytes) -> tuple[bytes, str]:
    """Convert DHL report CSV → Amazon tracking .txt. Returns (txt_bytes, log_str)."""
    from datetime import datetime as _dt

    log = io.StringIO()

    df = pd.read_csv(
        io.BytesIO(csv_bytes),
        dtype={"Shipment number": str, "Alternative reference": str},
    )

    required = {"Customer reference", "Shipment number", "Alternative reference"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"DHL report is missing columns: {missing}")

    df = df[["Customer reference", "Shipment number", "Alternative reference"]].copy()
    log.write(f"Loaded {len(df)} row(s) from DHL report.\n\n")

    # ── Filter to today only ──────────────────────────────────────────────────
    today_dt = _dt.today()

    def _parse(val):
        for fmt in ("%d-%b", "%d-%b-%y", "%d-%b-%Y", "%d/%m/%Y", "%d/%m/%y"):
            try:
                return _dt.strptime(str(val).strip(), fmt)
            except ValueError:
                continue
        return None

    def _is_today(d):
        if d is None:
            return False
        if d.year == 1900:   # parsed without year e.g. "25-May"
            return d.day == today_dt.day and d.month == today_dt.month
        return d.date() == today_dt.date()

    today_mask = df["Alternative reference"].apply(_parse).apply(_is_today)
    dropped = (~today_mask).sum()

    if dropped:
        sample = ", ".join(df.loc[~today_mask, "Alternative reference"].astype(str).unique()[:10])
        log.write(f"Dropped {dropped} row(s) not dated today ({today_dt.strftime('%d-%b')}).\n")
        log.write(f"Dates found: {sample}\n\n")

    df = df[today_mask].copy()

    if df.empty:
        log.write(f"No rows for today ({today_dt.strftime('%d-%b')}) — output will have headers only.\n")
    else:
        log.write(f"{len(df)} row(s) for today — building Amazon tracking file.\n")

    # ── Build Amazon DataFrame ────────────────────────────────────────────────
    AMAZON_COLUMNS = [
        "order-id", "order-item-id", "quantity", "ship-date",
        "carrier-code", "carrier-name", "tracking-number", "ship-method",
        "transparency_code", "ship_from_address_name", "ship_from_address_line1",
        "ship_from_address_line2", "ship_from_address_line3", "ship_from_address_city",
        "ship_from_address_county", "ship_from_address_state_or_region",
        "ship_from_address_postalcode", "ship_from_address_countrycode",
    ]
    today_amz = f"{today_dt.month}/{today_dt.day}/{today_dt.year}"
    out = pd.DataFrame(index=df.index, columns=AMAZON_COLUMNS)
    out["order-id"]        = df["Customer reference"].str.strip()
    out["tracking-number"] = df["Shipment number"].astype(str).str.strip()
    out["ship-date"]       = today_amz
    out["carrier-code"]    = "DHL eCommerce"
    out["ship-method"]     = "Premier 24"
    out = out.fillna("").reset_index(drop=True)

    # ── Export to bytes ───────────────────────────────────────────────────────
    buf = io.StringIO()
    out.to_csv(buf, sep="\t", index=False, lineterminator="\r\n")
    log.write(f"\nDone. {len(out)} row(s) written.")
    return buf.getvalue().encode("utf-8"), log.getvalue()


# ─── Sidebar ──────────────────────────────────────────────────────────────────
with st.sidebar:
    st.title("Settings")

    # Non-shipment tracker status
    st.subheader("Non-Shipment Tracker")
    _src = st.session_state.ns_source
    if _src == "github":
        st.success("Loaded from GitHub (auto-syncs after each run)")
    elif _src == "local":
        st.info("Loaded from bundled file (local/no GitHub secrets set)")
    elif _src == "github_error":
        st.warning("Could not reach GitHub — upload the file manually below.")
        _ns_ov = st.file_uploader("Non-shipment XLSX", type=["xlsx", "xls"], key="ns_override")
        if _ns_ov:
            st.session_state.non_shipment_bytes = _ns_ov.getvalue()
            st.session_state.ns_source = "manual"
    else:
        st.warning("Non-shipment tracker not found — filtering previous orders will be skipped.")
        _ns_ov = st.file_uploader("Upload non-shipment XLSX (optional)", type=["xlsx", "xls"], key="ns_override")
        if _ns_ov:
            st.session_state.non_shipment_bytes = _ns_ov.getvalue()
            st.session_state.ns_source = "manual"

    # Reference file status
    st.subheader("Reference Files")
    for _label, _attr, _default in [
        ("PNC/BTC Weights", "pnc_btc_bytes", _DEFAULT_PNC),
        ("DFF Weights",     "dff_bytes",     _DEFAULT_DFF),
        ("D-Zone Reference","dzone_bytes",   _DEFAULT_DZONE),
    ]:
        if st.session_state[_attr]:
            st.success(f"{_label}: bundled")
        else:
            st.warning(f"{_label}: not found")

    st.divider()
    st.subheader("AI Settings (optional)")
    _ai_key = st.text_input(
        "OpenAI API Key",
        type="password",
        value=st.session_state.openai_key,
        help="Enables GPT-powered delivery instruction trimming. Leave blank for rule-based trimming.",
    )
    if _ai_key != st.session_state.openai_key:
        st.session_state.openai_key = _ai_key
    if _ai_key:
        st.session_state.use_ai = st.toggle("Enable AI trimming", value=True)

    st.divider()
    if st.button("Reset / Start Over", use_container_width=True):
        _reset()
        st.rerun()


# ─── Main content ─────────────────────────────────────────────────────────────
st.title("DHL Label Creator")

tab1, tab2 = st.tabs(["Label Creation", "Amazon Tracking Update"])


# ══════════════════════════════════════════════════════════════════════════════
# Tab 1 — Label Creation Pipeline
# ══════════════════════════════════════════════════════════════════════════════
with tab1:
    st.caption("Amazon unshipped orders (.txt) → DHL import CSV")

    if st.session_state.error:
        st.error(f"**Error:** {st.session_state.error}")
        if st.button("Dismiss"):
            st.session_state.error = None
            st.rerun()

    elif st.session_state.stage == "upload":
        st.subheader("Upload Amazon Unshipped Orders")
        st.markdown(
            "Export the **unshipped orders** report from Amazon Seller Central "
            "(`Orders → Manage Orders → Unshipped → Export`), then upload the `.txt` file."
        )

        amazon_file = st.file_uploader(
            "Amazon unshipped orders (.txt, tab-delimited)",
            type=["txt"],
            key="amazon_uploader",
        )

        if amazon_file:
            st.info(f"File ready: **{amazon_file.name}** ({len(amazon_file.getvalue()) // 1024} KB)")

            if st.button("Run Pre-Processing (Steps 1–3)", type="primary"):
                with st.spinner("Ingesting, filtering existing orders, and parsing SKUs…"):
                    try:
                        state, logs = run_phase1(amazon_file.getvalue())
                        st.session_state.pipeline_state = state
                        st.session_state.logs_phase1    = logs
                        st.session_state.stage          = "sku_review"
                        st.rerun()
                    except Exception as exc:
                        st.session_state.error = str(exc)
                        st.rerun()

    elif st.session_state.stage == "sku_review":
        state = st.session_state.pipeline_state

        with st.expander("Pre-processing log (steps 1–3)", expanded=False):
            st.code(st.session_state.logs_phase1, language=None)

        st.subheader("SKU Review")
        st.markdown(
            f"**{len(state.df)} orders** ready. "
            "Tick **Remove** on any order that should NOT be labeled, then click Approve."
        )

        review_cols = [c for c in ["order-id", "sku", "supplier", "uid", "brand", "color", "size", "pack"]
                       if c in state.df.columns]
        review_df = state.df[review_cols].copy()
        review_df.insert(0, "Remove", False)

        edited = st.data_editor(
            review_df,
            use_container_width=True,
            hide_index=True,
            column_config={
                "Remove": st.column_config.CheckboxColumn(
                    "Remove",
                    help="Tick to exclude this order",
                    default=False,
                ),
            },
        )

        flagged_oids = (
            edited.loc[edited["Remove"].astype(bool), "order-id"].tolist()
            if "order-id" in edited.columns else []
        )
        if flagged_oids:
            st.warning(f"**{len(flagged_oids)} order(s) will be removed:** {', '.join(str(o) for o in flagged_oids)}")

        col_ok, col_back = st.columns([3, 1])

        with col_ok:
            if st.button("Approve & Generate Labels", type="primary", use_container_width=True):
                rows_in_gate = len(state.df)
                if flagged_oids:
                    mask = state.df["order-id"].isin(flagged_oids)
                    state.df = state.df[~mask].copy().reset_index(drop=True)

                if len(state.df) == 0:
                    st.session_state.error = "No orders remain after SKU review — nothing to label."
                    st.session_state.stage = "upload"
                    st.rerun()

                state.log(
                    step="step4_sku_gate",
                    rows_in=rows_in_gate,
                    note=f"Approved via web UI; removed {len(flagged_oids)} order(s)",
                )
                st.session_state.pipeline_state = state

                with st.spinner("Running steps 5–11 (weights, D-zone, address, AI trim, export)…"):
                    try:
                        outputs, logs = run_phase2(state, skipped_oids=flagged_oids)
                        st.session_state.out_dhl     = outputs.get("dhl")
                        st.session_state.out_audit   = outputs.get("audit")
                        st.session_state.out_dzone   = outputs.get("dzone")
                        st.session_state.out_nonship = outputs.get("non_shipment")
                        st.session_state.logs_phase2 = logs

                        # ── Auto-commit updated non-shipment tracker to GitHub ──
                        token, repo = _gh_secrets()
                        if (token and repo
                                and st.session_state.ns_sha
                                and st.session_state.out_nonship):
                            today_str = date.today().strftime("%d-%b-%Y")
                            new_sha = _gh_put(
                                token, repo,
                                _GH_NS_PATH,
                                st.session_state.out_nonship,
                                st.session_state.ns_sha,
                                f"Update non-shipment tracker {today_str}",
                            )
                            if new_sha:
                                st.session_state.ns_sha             = new_sha
                                st.session_state.non_shipment_bytes = st.session_state.out_nonship
                                st.session_state.gh_update_ok       = True
                            else:
                                st.session_state.gh_update_ok = False
                        else:
                            st.session_state.gh_update_ok = None

                        st.session_state.stage = "complete"
                        st.rerun()
                    except Exception as exc:
                        st.session_state.error = str(exc)
                        st.rerun()

        with col_back:
            if st.button("Back / Re-upload", use_container_width=True):
                _reset()
                st.rerun()

    elif st.session_state.stage == "complete":
        st.success("Pipeline complete!")

        if st.session_state.gh_update_ok is True:
            st.info("Non-shipment tracker updated in GitHub — next run will see today's orders.")
        elif st.session_state.gh_update_ok is False:
            st.warning(
                "Could not auto-update non-shipment tracker on GitHub. "
                "Download the updated file below and re-upload it next session."
            )

        today = date.today().strftime("%Y-%m-%d")
        left, right = st.columns(2)

        with left:
            if st.session_state.out_dhl:
                st.download_button(
                    "Download DHL Import CSV",
                    data=st.session_state.out_dhl,
                    file_name=f"dhl_import_{today}.csv",
                    mime="text/csv",
                    use_container_width=True,
                    type="primary",
                )
            else:
                st.warning("DHL import file was not generated.")

            if st.session_state.out_audit:
                st.download_button(
                    "Download Audit Log",
                    data=st.session_state.out_audit,
                    file_name=f"audit_log_{today}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

        with right:
            if st.session_state.out_nonship:
                st.download_button(
                    "Download Updated Non-Shipment Tracker",
                    data=st.session_state.out_nonship,
                    file_name=f"non-shipment-updated-{today}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    help="Only needed if GitHub auto-sync failed.",
                )

            if st.session_state.out_dzone:
                st.download_button(
                    "Download Dropped D-Zone Orders",
                    data=st.session_state.out_dzone,
                    file_name=f"dzone_dropped_{today}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

        with st.expander("Full pipeline log", expanded=False):
            st.code(
                st.session_state.logs_phase1 + "\n" + st.session_state.logs_phase2,
                language=None,
            )

        st.divider()
        if st.button("Process Another Batch", type="secondary"):
            _reset()
            st.rerun()


# ══════════════════════════════════════════════════════════════════════════════
# Tab 2 — Amazon Tracking Update
# ══════════════════════════════════════════════════════════════════════════════
with tab2:
    st.caption("DHL report CSV → Amazon tracking update .txt")
    st.subheader("Upload DHL Report")
    st.markdown(
        "Download the shipment report from DHL, then upload the CSV here. "
        "The app filters to **today's rows only** and produces the Amazon tracking file."
    )

    dhl_csv = st.file_uploader(
        "DHL report (.csv)",
        type=["csv"],
        key="tracking_uploader",
    )

    if dhl_csv:
        st.info(f"File ready: **{dhl_csv.name}**")

        if st.button("Generate Amazon Tracking File", type="primary"):
            with st.spinner("Processing…"):
                try:
                    txt_bytes, log = _run_tracking(dhl_csv.getvalue())
                    st.session_state.tracking_out = txt_bytes
                    st.session_state.tracking_log = log
                except Exception as exc:
                    st.error(f"**Error:** {exc}")

    if st.session_state.tracking_out:
        today = date.today().strftime("%Y-%m-%d")
        st.download_button(
            "Download Amazon Tracking .txt",
            data=st.session_state.tracking_out,
            file_name=f"Amazon_Trackings_Update_DHL_{today}.txt",
            mime="text/plain",
            use_container_width=True,
            type="primary",
        )

        with st.expander("Processing log", expanded=True):
            st.code(st.session_state.tracking_log, language=None)

        if st.button("Clear / New File"):
            st.session_state.tracking_out = None
            st.session_state.tracking_log = ""
            st.rerun()
